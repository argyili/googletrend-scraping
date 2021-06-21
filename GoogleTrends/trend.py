# from .pytrends.request import TrendReq
from pytrends.request import TrendReq
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import rcParams
import openpyxl
from pathlib import Path

def fecth_list_from_xlsx():
    xlsx_file = Path('geo_code_jp.xlsx')
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    tmp = '0'
    column_num = []
    for col in ws.iter_cols(min_row=2, max_col=1, values_only=True):
        for cell in col:
            if cell != tmp :
                column_num.append(cell)
                tmp = cell
    
    column_alpha = []
    for col in ws.iter_cols(min_row=2, min_col=4, max_col=4, values_only=True):
        for cell in col:
            if cell != tmp:
                column_alpha.append(cell)
                tmp = cell
    matrix = list(zip(column_num, column_alpha))
    # for i in range(0, len(column_num)):
    #     matrix.append(column_num[i], column_alpha[i])
    return matrix

def interest_over_time(kw_list, timeframe, geo, geo_name):
    pytrend.build_payload(kw_list, timeframe=timeframe, geo=geo, geo_name=geo_name)
    df = pytrend.interest_over_time()
    return df

def interest_by_region(kw_list, timeframe, geo, geo_name):
    pytrend.build_payload(kw_list, timeframe=timeframe, geo=geo, geo_name=geo_name)
    df = pytrend.interest_by_region(inc_low_vol=True, inc_geo_code=False)
    return df

def fetch_interest(kw_list, timeframe, matrix):
    df_sum = None
    for i in range(47):
        index = str(i + 1)
        geo_name = matrix[i][1]
        if (i + 1) < 10:
            index = '0' + index
        df = interest_over_time(kw_list, timeframe, geo='JP-' + index, geo_name=geo_name)
        df_sum = pd.concat([df_sum, df], axis=1)

    df_sum.to_pickle('./files/df_sum_' + timeframe + '.pkl')
    output  = pd.read_pickle('./files/df_sum_' + timeframe + '.pkl')
    plt.rcParams['font.family'] = 'IPAexGothic'
    output.plot(figsize=(20,20))
    plt.savefig('./files/pollen_japan_' + timeframe + '.jpg')

    with pd.ExcelWriter('./files/pollen_japan_' + timeframe + '.xlsx', date_format='YYYY-MM-DD') as writer:
        output.to_excel(writer)
    
    fetch_interestValue_regions(kw_list, timeframe, matrix)

def fetch_interestValue_regions(kw_list, timeframe, matrix):
    df_regions_ = interest_by_region(kw_list, timeframe, geo='JP', geo_name='Japan')
    df_regions_.to_pickle('./files/df_regions_' + timeframe + '.pkl')
    output  = pd.read_pickle('./files/df_regions_' + timeframe + '.pkl')
    plt.rcParams['font.family'] = 'IPAexGothic'
    output.plot(figsize=(20,20))
    plt.savefig('./files/pollen_japan_regions_' + timeframe + '.jpg')

    with pd.ExcelWriter('./files/pollen_japan_regions_' + timeframe + '.xlsx', date_format='YYYY-MM-DD') as writer:
        output.to_excel(writer)

def fetch_interestValue_cities(kw_list, timeframe, matrix):
    df_sum = None
    for i in range(47):
        index = str(i + 1)
        geo_name = matrix[i][1]
        if (i + 1) < 10:
            index = '0' + index
        df = interest_by_region(kw_list, timeframe, geo='JP-' + index, geo_name=geo_name)
        df_sum = pd.concat([df_sum, df], axis=1)

    df_sum = pd.concat([df_sum, df], axis=1)
    df_sum.to_pickle('./files/df_sum_cities_' + timeframe + '.pkl')

    output  = pd.read_pickle('./files/df_sum_cities_' + timeframe + '.pkl')
    plt.rcParams['font.family'] = 'IPAexGothic'
    output.plot(figsize=(20,20))
    plt.savefig('./files/pollen_japan_cities_' + timeframe + '.jpg')

    with pd.ExcelWriter('./files/pollen_japan_cities_' + timeframe + '.xlsx', date_format='YYYY-MM-DD') as writer:
        output.to_excel(writer)

if __name__ == '__main__':
    
    pytrend = TrendReq(hl='ja-jp',tz=540)
    kw_list = ['花粉症']
    matrix = fecth_list_from_xlsx()

    # Automatically fetch daily data for short period
    timeframe='2018-02-01 2018-06-30'
    fetch_interest(kw_list, timeframe, matrix)
    fetch_interestValue_cities(kw_list, timeframe, matrix)

    timeframe='2019-02-01 2019-06-30'
    fetch_interest(kw_list, timeframe, matrix)
    fetch_interestValue_cities(kw_list, timeframe, matrix)

    # Automatically fetch weekly data for long period
    timeframe='2018-02-01 2019-06-30'
    fetch_interest(kw_list, timeframe, matrix)
    fetch_interestValue_cities(kw_list, timeframe, matrix)

    # Automatically fetch monthly data for even longer period
    timeframe='2004-01-01 2019-06-30'
    fetch_interest(kw_list, timeframe, matrix)
    fetch_interestValue_cities(kw_list, timeframe, matrix)
